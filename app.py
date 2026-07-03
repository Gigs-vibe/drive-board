# -*- coding: utf-8 -*-
"""
Taska — доска задач (десктоп-версия для Windows 10/11)

Запускает board.html в нативном окне, держит иконку в трее и присылает
системные уведомления о задачах со сроком — даже когда окно свёрнуто в трей.

Зависимости: см. requirements.txt
Сборка в .exe: см. README.txt
"""
import truststore
truststore.inject_into_ssl()

# Явно указываем .NET Framework до импорта webview — иначе pythonnet может не найти runtime
try:
    import pythonnet
    pythonnet.load("netfx")
except Exception:
    pass

import os
import sys
import json
import time
import threading
import urllib.request
import urllib.parse
import secrets
import hashlib
import base64
import http.server
import webbrowser

import socket
import webview  # pywebview

# --- уведомления и трей только под Windows ---
try:
    from winotify import Notification, audio
    HAVE_NOTIFY = True
except Exception:
    HAVE_NOTIFY = False

try:
    import pystray
    from PIL import Image
    HAVE_TRAY = True
except Exception:
    HAVE_TRAY = False


APP_NAME = "Taska — доска задач"
window = None  # ссылка на окно pywebview (нужна для диалога сохранения файла)

# === Версия и обновления через GitHub ===
# При каждом новом релизе увеличивай VERSION и ставь такой же тег у релиза (например v1.1).
VERSION = "1.6.3"
GITHUB_REPO = "Gigs-vibe/drive-board"
SINGLE_INSTANCE_PORT = 27315  # локальный порт для обнаружения запущенного экземпляра

_pending_update = {"version": None, "download_url": None}

# === Supabase (для входа через Google) ===
SUPABASE_URL = "https://qbwjnkwzispjbarrlgkf.supabase.co"
SUPABASE_ANON = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InFid2pua3d6aXNwamJhcnJsZ2tmIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODI0NjEwMTEsImV4cCI6MjA5ODAzNzAxMX0.GBoFN3pcyNDC2jZbhcb3VClY7TU-BGn0d2q1PYXkXwE"
GOOGLE_PORT = 8765  # на этот адрес ловим ответ; он должен быть в Supabase → URL Configuration → Redirect URLs как http://localhost:8765/

# ----------------------------------------------------------------------------
# Пути к файлам
# ----------------------------------------------------------------------------
def resource_path(name: str) -> str:
    """Файл, упакованный внутрь .exe (board.html, icon.*)."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, name)


def data_dir() -> str:
    """Папка с данными доски: %APPDATA%\\Taska для установленной версии, иначе рядом со скриптом."""
    if getattr(sys, "frozen", False):
        base = os.path.join(os.environ.get("APPDATA", os.path.dirname(sys.executable)), "Taska")
        os.makedirs(base, exist_ok=True)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return base


DATA_FILE = os.path.join(data_dir(), "drive-board.json")      # данные доски (пишет JS)
NOTIFIED_FILE = os.path.join(data_dir(), "drive-notified.json")  # что уже напомнили (пишет Python)


# ----------------------------------------------------------------------------
# Мост JS <-> Python: сохранение и загрузка доски
# ----------------------------------------------------------------------------
class Api:
    def get_version(self):
        return VERSION

    def load(self):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                return f.read()
        except FileNotFoundError:
            return ""
        except Exception:
            return ""

    def save(self, data):
        try:
            with open(DATA_FILE, "w", encoding="utf-8") as f:
                f.write(data)
            return True
        except Exception:
            return False

    def export_excel(self, data_json):
        """Строит .xlsx с двумя листами: задачи и журнал смен статусов."""
        try:
            import openpyxl
            from openpyxl.styles import Font
        except Exception:
            return {"ok": False, "error": "Не установлен openpyxl (см. requirements.txt)"}
        try:
            data = json.loads(data_json)
        except Exception:
            return {"ok": False, "error": "Не удалось прочитать данные доски"}

        RU = {"high": "Высокий", "med": "Средний", "low": "Низкий", "": "—", None: "—"}
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Задачи"
        ws1.append(["Колонка", "Задача", "Приоритет", "Дедлайн", "Создана"])
        for col in data.get("columns", []):
            for card in col.get("cards", []):
                if card.get("repeat") == "daily":
                    dl = "каждый день " + (card.get("dueTime") or "")
                else:
                    dl = card.get("due") or ""
                created = ""
                if card.get("created"):
                    created = time.strftime("%d.%m.%Y %H:%M", time.localtime(card["created"] / 1000))
                ws1.append([col.get("title", ""), card.get("title", ""),
                            RU.get(card.get("prio"), "—"), dl, created])

        ws2 = wb.create_sheet("История статусов")
        ws2.append(["Задача", "Из", "В", "Дата", "Время"])
        for e in data.get("log", []):
            ts = e.get("ts")
            d_str = t_str = ""
            if ts:
                lt = time.localtime(ts / 1000)
                d_str = time.strftime("%d.%m.%Y", lt)
                t_str = time.strftime("%H:%M:%S", lt)
            ws2.append([e.get("title", ""), e.get("from", ""), e.get("to", ""), d_str, t_str])

        for ws in (ws1, ws2):
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.column_dimensions["A"].width = 26
            ws.column_dimensions["B"].width = 30

        try:
            path = window.create_file_dialog(webview.SAVE_DIALOG,
                                             save_filename="Доска.xlsx",
                                             file_types=("Excel (*.xlsx)",))
        except Exception as ex:
            return {"ok": False, "error": str(ex)}
        if not path:
            return {"ok": False, "cancelled": True}
        if isinstance(path, (list, tuple)):
            path = path[0]
        path = str(path)
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            wb.save(path)
        except Exception as ex:
            return {"ok": False, "error": str(ex)}
        return {"ok": True, "path": path}

    def download_and_update(self, download_url):
        """Скачивает TaskaSetup.exe и запускает его тихо через ShellExecuteW (переживает закрытие приложения)."""
        if not getattr(sys, "frozen", False):
            webbrowser.open(download_url)
            return {"ok": False, "error": "dev-режим: открыт браузер"}

        staging = os.path.join(os.environ.get("TEMP", r"C:\Temp"), "Taska")
        os.makedirs(staging, exist_ok=True)
        installer = os.path.join(staging, "TaskaSetup.exe")

        try:
            def on_progress(count, block, total):
                if total > 0 and window:
                    pct = min(int(count * block * 100 / total), 99)
                    window.evaluate_js(f"updateDownloadProgress({pct})")
            urllib.request.urlretrieve(download_url, installer, reporthook=on_progress)
        except Exception as ex:
            return {"ok": False, "error": f"Ошибка загрузки: {ex}"}

        # Проверяем что скачали настоящий установщик, а не HTML-страницу
        try:
            size = os.path.getsize(installer)
            with open(installer, "rb") as f:
                magic = f.read(2)
            if magic != b"MZ" or size < 1_000_000:
                return {"ok": False,
                        "error": f"Скачан не установщик (размер {size} б). Скачай вручную с GitHub."}
        except Exception as ex:
            return {"ok": False, "error": f"Ошибка проверки: {ex}"}

        # Снимаем отметку «загружен из интернета» — иначе SmartScreen блокирует запуск
        try:
            with open(installer + ":Zone.Identifier", "w") as f:
                f.write("[ZoneTransfer]\nZoneId=0\n")
        except Exception:
            pass

        # Bat-обёртка: ждём 3 сек (приложение успеет закрыться по os._exit), ставим, перезапуск делает сам установщик
        wrapper = os.path.join(staging, "run_update.bat")
        try:
            with open(wrapper, "w", encoding="ascii") as f:
                f.write(
                    "@echo off\r\n"
                    "timeout /t 3 /nobreak >nul\r\n"
                    f'"{installer}" /VERYSILENT /NORESTART\r\n'
                    'del /f /q "%~f0"\r\n'
                )
        except Exception as ex:
            return {"ok": False, "error": f"Ошибка скрипта: {ex}"}

        # ShellExecuteW запускает bat через оболочку — процесс независим от нас и переживёт os._exit(0).
        # SW_HIDE (0) прячет окно cmd. Прямой запуск установщика проверен — работает надёжнее Task Scheduler.
        import ctypes
        try:
            r = ctypes.windll.shell32.ShellExecuteW(None, "open", wrapper, None, staging, 0)
            if r <= 32:
                return {"ok": False, "error": f"Не удалось запустить установку (код {r})"}
        except Exception as ex:
            return {"ok": False, "error": f"Ошибка запуска: {ex}"}

        time.sleep(1)
        os._exit(0)

    def google_login(self):
        """Вход через Google: открывает браузер, ловит ответ на localhost, меняет код на сессию (PKCE)."""
        def b64url(b):
            return base64.urlsafe_b64encode(b).decode().rstrip("=")
        verifier = b64url(secrets.token_bytes(48))
        challenge = b64url(hashlib.sha256(verifier.encode()).digest())
        redirect = "http://localhost:%d/" % GOOGLE_PORT
        authorize = (SUPABASE_URL + "/auth/v1/authorize?provider=google"
                     + "&redirect_to=" + urllib.parse.quote(redirect, safe="")
                     + "&code_challenge=" + challenge + "&code_challenge_method=s256")
        result = {}

        class Handler(http.server.BaseHTTPRequestHandler):
            def log_message(self, *a):
                pass
            def do_GET(self):
                params = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
                got = False
                if "code" in params:
                    result["code"] = params["code"][0]; got = True
                elif "error_description" in params or "error" in params:
                    result["error"] = (params.get("error_description") or params.get("error"))[0]; got = True
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                if got:
                    self.wfile.write(("<html><body style='font-family:Segoe UI,sans-serif;"
                                      "text-align:center;padding-top:80px;color:#171a21'>"
                                      "<h2>Готово!</h2><p>Можно вернуться в приложение «Taska».</p>"
                                      "</body></html>").encode("utf-8"))
                else:
                    self.wfile.write(b"")

        try:
            httpd = http.server.HTTPServer(("localhost", GOOGLE_PORT), Handler)
        except Exception as ex:
            return {"ok": False, "error": "Порт %d занят (%s). Закрой лишние программы и попробуй снова." % (GOOGLE_PORT, ex)}
        httpd.timeout = 2
        webbrowser.open(authorize)
        start = time.time()
        while not result and time.time() - start < 180:
            httpd.handle_request()
        httpd.server_close()

        if result.get("error"):
            return {"ok": False, "error": result["error"]}
        code = result.get("code")
        if not code:
            return {"ok": False, "error": "Не дождались ответа от Google (таймаут)."}
        try:
            body = json.dumps({"auth_code": code, "code_verifier": verifier}).encode()
            req = urllib.request.Request(
                SUPABASE_URL + "/auth/v1/token?grant_type=pkce",
                data=body,
                headers={"apikey": SUPABASE_ANON, "Content-Type": "application/json"},
                method="POST")
            with urllib.request.urlopen(req, timeout=15) as r:
                data = json.load(r)
            if not data.get("access_token"):
                return {"ok": False, "error": "Сервер не вернул сессию."}
            return {"ok": True, "session": data}
        except Exception as ex:
            return {"ok": False, "error": str(ex)}


# ----------------------------------------------------------------------------
# Фоновые напоминания
# ----------------------------------------------------------------------------
def load_notified() -> set:
    try:
        with open(NOTIFIED_FILE, "r", encoding="utf-8") as f:
            return set(json.load(f))
    except Exception:
        return set()


def save_notified(s: set):
    try:
        with open(NOTIFIED_FILE, "w", encoding="utf-8") as f:
            json.dump(list(s), f)
    except Exception:
        pass


def show_toast(title: str, msg: str):
    if not HAVE_NOTIFY:
        return
    try:
        icon = resource_path("icon.png")
        t = Notification(app_id=APP_NAME, title=title, msg=msg,
                         icon=icon if os.path.exists(icon) else "")
        t.set_audio(audio.Default, loop=False)
        t.show()
    except Exception:
        pass


# ----------------------------------------------------------------------------
# Проверка обновлений на GitHub
# ----------------------------------------------------------------------------
def parse_ver(s):
    """'v1.2.3' -> (1, 2, 3) для сравнения версий."""
    s = (s or "").lstrip("vV").strip()
    parts = []
    for p in s.split("."):
        digits = "".join(ch for ch in p if ch.isdigit())
        parts.append(int(digits) if digits else 0)
    return tuple(parts) if parts else (0,)


def get_asset_url(release_data):
    """Возвращает прямую ссылку на установщик (TaskaSetup.exe) из активов релиза."""
    for asset in release_data.get("assets", []):
        if asset.get("name", "").lower() == "taskasetup.exe":
            return asset.get("browser_download_url")
    return None


def show_update_toast(latest, page_url):
    if not HAVE_NOTIFY:
        return
    try:
        icon = resource_path("icon.png")
        t = Notification(app_id=APP_NAME, title="Доступно обновление",
                         msg=f"Вышла версия {latest} (у тебя {VERSION}). Скачай новый Taska.exe с GitHub.",
                         icon=icon if os.path.exists(icon) else "")
        t.add_actions(label="Скачать", launch=page_url)
        t.set_audio(audio.Default, loop=False)
        t.show()
    except Exception:
        pass


def check_update(manual=False):
    """Сравнивает VERSION с последним релизом на GitHub и уведомляет, если есть новее."""
    global _pending_update
    if not GITHUB_REPO:
        if manual:
            show_toast("Обновления", "Адрес репозитория не задан в app.py (GITHUB_REPO).")
        return
    try:
        api_url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
        req = urllib.request.Request(api_url, headers={"User-Agent": "Taska-Board"})
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.load(r)
        latest = data.get("tag_name", "")
        page = data.get("html_url", f"https://github.com/{GITHUB_REPO}/releases/latest")
        download_url = get_asset_url(data) or page
        if parse_ver(latest) > parse_ver(VERSION):
            _pending_update["version"] = latest
            _pending_update["download_url"] = download_url
            show_update_toast(latest, page)
            if window:
                safe_url = download_url.replace("\\", "\\\\").replace("'", "\\'")
                window.evaluate_js(f"showUpdateBanner('{latest}', '{safe_url}')")
        elif manual:
            show_toast("Обновлений нет", f"У тебя последняя версия ({VERSION}).")
            if window:
                window.evaluate_js(f"showToast('Обновлений нет', 'У тебя последняя версия ({VERSION}).', '')")
    except Exception:
        if manual:
            show_toast("Обновления", "Не удалось проверить — нет интернета или репозиторий недоступен.")


def reminder_loop(stop_event: threading.Event):
    """Каждые 15 секунд читаем доску и шлём пуши: за 30 минут до дедлайна и в дедлайн."""
    PRE = 30 * 60  # за сколько секунд предупредить заранее
    notified = load_notified()
    while not stop_event.is_set():
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                state = json.load(f)
            now = time.time()
            changed = False
            for col in state.get("columns", []):
                if col.get("role") == "done":
                    continue
                col_title = col.get("title", "")
                for card in col.get("cards", []):
                    cid = card.get("id")
                    title = card.get("title", "Задача")

                    # повтор «каждый день» в заданное время
                    if card.get("repeat") == "daily" and card.get("dueTime"):
                        try:
                            hh, mm = card["dueTime"].split(":")
                            lt = time.localtime(now)
                            sched = time.mktime((lt.tm_year, lt.tm_mon, lt.tm_mday,
                                                 int(hh), int(mm), 0, 0, 0, -1))
                        except Exception:
                            sched = None
                        if sched is not None:
                            day_key = f"{cid}|daily|{lt.tm_year}-{lt.tm_mon}-{lt.tm_mday}"
                            if now >= sched and day_key not in notified:
                                show_toast("🔔 Ежедневно: " + title, "Время задачи — " + col_title)
                                notified.add(day_key); changed = True
                        continue

                    due = card.get("due")
                    if not due:
                        continue
                    try:
                        # формат datetime-local: "2026-06-25T14:30"
                        due_ts = time.mktime(time.strptime(due[:16], "%Y-%m-%dT%H:%M"))
                    except Exception:
                        continue
                    # за 30 минут (только если мы реально в окне "до дедлайна")
                    pre_key = f"{cid}|{due}|pre"
                    if pre_key not in notified and due_ts - PRE <= now < due_ts:
                        show_toast("⏰ Через 30 минут: " + title, "Дедлайн скоро — " + col_title)
                        notified.add(pre_key); changed = True
                    # в момент дедлайна
                    at_key = f"{cid}|{due}|at"
                    if at_key not in notified and now >= due_ts:
                        show_toast("⏰ Дедлайн: " + title, "Срок наступил — " + col_title)
                        notified.add(at_key); changed = True
            if changed:
                save_notified(notified)
        except FileNotFoundError:
            pass
        except Exception:
            pass
        stop_event.wait(15)


# ----------------------------------------------------------------------------
# Трей
# ----------------------------------------------------------------------------
def build_tray(window, stop_event):
    if not HAVE_TRAY:
        return None
    try:
        img = Image.open(resource_path("icon.png"))
    except Exception:
        img = Image.new("RGB", (64, 64), (244, 99, 58))

    def do_open(icon, item):
        try:
            window.show()
        except Exception:
            pass

    def do_quit(icon, item):
        stop_event.set()
        try:
            icon.stop()
        except Exception:
            pass
        try:
            window.destroy()
        except Exception:
            pass

    def do_check(icon, item):
        threading.Thread(target=check_update, args=(True,), daemon=True).start()

    menu = pystray.Menu(
        pystray.MenuItem("Открыть", do_open, default=True),
        pystray.MenuItem("Проверить обновления", do_check),
        pystray.MenuItem("Выход", do_quit),
    )
    return pystray.Icon("drive", img, f"{APP_NAME} v{VERSION}", menu)


# ----------------------------------------------------------------------------
# Single-instance: второй запуск показывает окно первого и закрывается
# ----------------------------------------------------------------------------
def try_bring_to_front() -> bool:
    """Пытается найти уже запущенный экземпляр и попросить его показать окно."""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(1)
        s.connect(("127.0.0.1", SINGLE_INSTANCE_PORT))
        s.sendall(b"show")
        s.close()
        return True
    except Exception:
        return False


def start_single_instance_server(stop_event: threading.Event):
    """Слушает порт; при команде 'show' выводит окно на передний план."""
    try:
        srv = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        srv.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        srv.bind(("127.0.0.1", SINGLE_INSTANCE_PORT))
        srv.listen(5)
        srv.settimeout(1)
        while not stop_event.is_set():
            try:
                conn, _ = srv.accept()
                data = conn.recv(64)
                conn.close()
                if b"show" in data and window:
                    window.show()
            except socket.timeout:
                continue
            except Exception:
                break
        srv.close()
    except Exception:
        pass


# ----------------------------------------------------------------------------
# Запуск
# ----------------------------------------------------------------------------
def main():
    global window

    # Если уже запущен — показываем существующее окно и выходим
    if try_bring_to_front():
        return

    stop_event = threading.Event()
    api = Api()

    window = webview.create_window(
        APP_NAME,
        resource_path("board.html"),
        js_api=api,
        width=1280, height=820,
        min_size=(720, 520),
    )

    # крестик окна сворачивает в трей, а не закрывает программу
    def on_closing():
        if HAVE_TRAY and not stop_event.is_set():
            window.hide()
            return False  # отменяем закрытие
        return True
    window.events.closing += on_closing

    tray = build_tray(window, stop_event)
    if tray is not None:
        threading.Thread(target=tray.run, daemon=True).start()

    threading.Thread(target=reminder_loop, args=(stop_event,), daemon=True).start()
    threading.Thread(target=start_single_instance_server, args=(stop_event,), daemon=True).start()

    # проверка обновлений на GitHub (тихо, в фоне, через несколько секунд после старта)
    def delayed_update_check():
        time.sleep(6)
        check_update(False)
    threading.Thread(target=delayed_update_check, daemon=True).start()

    # WebView2 (Edge) встроен в Windows 10/11 — отдельной установки не нужно
    # storage_path + private_mode=False — localStorage сохраняется между запусками
    webview.start(storage_path=data_dir(), private_mode=False, gui="edgechromium")
    stop_event.set()


if __name__ == "__main__":
    main()
